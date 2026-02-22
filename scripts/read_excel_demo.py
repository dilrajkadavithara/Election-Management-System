import openpyxl
import os

def read_excel(file_path):
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    print(f"Reading sheet: {sheet.title}")
    
    # Read headers
    headers = [cell.value for cell in sheet[1]]
    print(f"Headers: {headers}")
    
    # Summarize first 10 rows
    data = []
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=11, values_only=True)):
        data.append(row)
    
    print("\nFirst 10 Data Rows:")
    for row in data:
        print(row)
        
    # Count total booths (assuming each row is a booth)
    total_booths = sheet.max_row - 1
    print(f"\nTotal Booths in Excel: {total_booths}")

if __name__ == "__main__":
    path = r"C:\Users\dilra\OneDrive\Desktop\Voterslist\Paravur_Polling_Stations_list (1).xlsx"
    read_excel(path)
